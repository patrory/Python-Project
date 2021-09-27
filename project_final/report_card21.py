from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
file_directory=input("enter your file location with file name  eg :C:\coding\contact.xlsx    :  ")
wb  = load_workbook(file_directory+'.xlsx')
ws = wb.active

print('''++++--------------------------------------------------------++++ 
                      1) Press A for making excel sheet
                      2) press B for attendance 
                      3) press C for adding ut marks
                      4) press D for enteringe endsem marks
                      4) press E for Total Attendance 
                      5) press F for report card
              -------------------------------------------------''')
  
task = input('which work you have to ')

if task=='A' or task=='a':
    print("excel sheet")
    working_day = int(input("enter the no of working days : "))
    work_day = []
    j=1
    for i in range(0,working_day) :
        work_day.append(j)
        j+=1
    q=1
    r=1
    for row in range(1,2):
    
        for col in range(7,working_day +7):
            char = get_column_letter(col)
            row1= get_column_letter(row)+char+str(row)
            if(col<=26):
                ws[char+str(row)].value=q
            if(col>26) :
              row1=get_column_letter(row)+get_column_letter(r) + str(row)
              ws[row1].value = q
              r+=1
            q+=1   
    ws['E1'].value= 'UT marks'
    ws['F1'].value='Endsem marks'         
    wb.save('contact.xlsx') 

elif task=='B' or task=='b':
    print("attendance monitoring")
    which_wkday=int(input("which is working day : "))
    fixedno=6+which_wkday

    rollno=[]
    for i in range(2,75):
      rollno.append(ws[f'A{i}'].value)

    print("add absent roll no")
    absent=[]
    while True :
        a=int(input('to add data in list press 1 , if not then 0 :  '))
        if a :
            b=int(input('roll no : '))
            absent.append(b)
        else :
            break 
    fixedno2=2
    for j in rollno:
        if j in absent:
            char=get_column_letter(fixedno)+str(fixedno2)
            ws[char].value=0
        else:
             char=get_column_letter(fixedno)+str(fixedno2)
             ws[char].value=1     
        fixedno2+=1
        wb.save('contact.xlsx')     

elif task=='c' or task=='C' :
        rollno=[]
        for i in range(2,75):
           rollno.append(ws[f'A{i}'].value)
        o=2   
        m=1
        for k in rollno :
            print(m)
            m+=1
            marks123=int(input("input marks of insem  ")) 
            ws['E'+str(o)].value=marks123 
            o+=1 
        wb.save('contact.xlsx')    

elif task=='d' or task=="D":
        rollno=[]
        for i in range(2,75):
           rollno.append(ws[f'A{i}'].value)
        o=2 
        m=1  
        for k in rollno :
            print(m)
            marks123=int(input("input marks of endsem")) 
            ws['F'+str(o)].value=marks123 
            o+=1 
            m+=1
        wb.save('contact.xlsx')     

elif task=='e'or task=='E':
    fixedno3=7
    count=0
    o=2
    r=1
    
    working_day=30   ## here ekk problem hai   problem how to access working day value from task== a or task==A
    rollno=[]
    for row in range(2,75):
        for col in range(7,working_day+7):
            if(col<=26):
               char=get_column_letter(fixedno3)+str(o)
               count +=int(ws[char].value)
               fixedno3+=1
            if(col>26):
                char='A'+get_column_letter(r)+str(o)
                count+=ws[char].value
                r+=1
                fixedno3+=1
        saved1='A'+get_column_letter(r+1)+str(o)     
        ws[saved1].value=count   
        print(count)

        o+=1
        fixedno3=7
        count=0
        r=1

    wb.save('contact.xlsx')    
elif task=='f' or task=='F' :
    print("report card")
    sum=0   
    for row in range(2,75):
        for col in range(5,7):
            char=get_column_letter(col)+str(row)   
           # print(char) 
            sum +=ws[char].value
            #print(sum)
        ws['AM'+str(row)].value=sum
        #print('AM'+str(row))
        sum=0
    wb.save('contact.xlsx')   
    print("welcome !  to PICT ")

    rollno_12=int(input('last 2 digit of roll no ::  '))
    rollno_12+=1
    print("whether you want total attendance or total marks scored in insem and insem  ")
    print('''1 for total attendance 
             2 for total marks   '''
         )
    abc123=int(input(' your answer : '))   
    if(abc123==1):{
        print(ws['AL'+str(rollno_12)].value)

    }  
    elif (abc123==2):{
        print(ws['AM'+str(rollno_12)].value )
    }
